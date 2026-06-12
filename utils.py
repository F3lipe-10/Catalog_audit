"""
================================================================================
MÓDULO DE UTILIDADES
================================================================================
Funciones para:
  - Leer archivos Excel (catálogo, BOT, asignaciones)
  - Procesar el catálogo aplicando todas las reglas
  - Exportar el resultado coloreado a Excel
================================================================================
"""

import io
import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import config
import clasificador
import ia_carne

# Subir el límite de celdas que Pandas Styler puede colorear.
# Por defecto son 262.144 celdas (~512x512). Si tu catálogo es más grande,
# levanta el límite aquí.
pd.set_option("styler.render.max_elements", 5_000_000)


# ------------------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------------------
def _normalizar_id(valor) -> str:
    """
    Convierte un ID a string de forma robusta:
      - 1103541.0   ->  "1103541"  (float de pandas con NaN en la columna)
      - "1103541.0" ->  "1103541"
      - "1103541 "  ->  "1103541"  (espacios)
      - NaN/None    ->  ""
      - "SYSCO123"  ->  "SYSCO123" (IDs no numéricos se devuelven limpios)
    """
    if pd.isna(valor):
        return ""
    s = str(valor).strip()
    try:
        return str(int(float(s)))
    except (ValueError, OverflowError):
        return s


# ------------------------------------------------------------------------------
# Lectura de archivos
# ------------------------------------------------------------------------------
@st.cache_data(show_spinner=False)
def leer_excel(archivo_bytes: bytes, nombre: str = "archivo", skip_rows: int = 0) -> pd.DataFrame:
    """Lee un archivo Excel desde bytes y devuelve un DataFrame.

    Args:
        archivo_bytes: contenido del Excel en bytes
        nombre: nombre descriptivo para mensajes de error
        skip_rows: filas a saltar antes del header (útil cuando el Excel
                   tiene filas de título/logo arriba de los nombres de columna)
    """
    try:
        df = pd.read_excel(io.BytesIO(archivo_bytes), skiprows=skip_rows)
        return df
    except Exception as e:
        st.error(f"❌ Error al leer {nombre}: {e}")
        return pd.DataFrame()


def cargar_asignaciones(ruta: str = None) -> pd.DataFrame:
    """
    Carga las asignaciones (distribuidor -> persona) desde la BD (Supabase).
    Si la BD falla o está vacía, intenta el Excel local como fallback.
    """
    import db
    cols_esperadas = [
        config.COL_ASSIGN_SUPPLIER_ID,
        config.COL_ASSIGN_SUPPLIER,
        config.COL_ASSIGN_PERSON,
    ]

    filas = db.cargar_asignaciones_db()
    if filas:
        df = pd.DataFrame(filas)
        df = df.rename(columns={
            "supplier_id":   config.COL_ASSIGN_SUPPLIER_ID,
            "supplier_name": config.COL_ASSIGN_SUPPLIER,
            "assigned_to":   config.COL_ASSIGN_PERSON,
        })
        return df[cols_esperadas]

    # Fallback: archivo Excel local
    ruta = ruta or config.RUTA_ASIGNACIONES
    try:
        df = pd.read_excel(ruta)
        faltantes = [c for c in cols_esperadas if c not in df.columns]
        if faltantes:
            st.warning(
                f"⚠️ El archivo de asignaciones no tiene estas columnas: "
                f"{', '.join(faltantes)}. Se usará vacío."
            )
            return pd.DataFrame(columns=cols_esperadas)
        return df
    except FileNotFoundError:
        st.warning("⚠️ No hay asignaciones en la BD ni archivo local.")
        return pd.DataFrame(columns=cols_esperadas)
    except Exception as e:
        st.error(f"❌ Error al leer asignaciones: {e}")
        return pd.DataFrame(columns=cols_esperadas)


def _normalizar_columnas_catalogo(df: pd.DataFrame) -> pd.DataFrame:
    """
    Renombra columnas del catálogo según ALIASES_COLUMNAS de config.py.
    La comparación es case-insensitive e ignora espacios extra.
    Devuelve el mismo DataFrame con las columnas renombradas.
    """
    aliases = getattr(config, "ALIASES_COLUMNAS", {})
    if not aliases:
        return df

    # Normalizar cada alias para comparar de forma flexible
    aliases_norm = {
        " ".join(k.upper().split()): v for k, v in aliases.items()
    }

    renombrar = {}
    for col in df.columns:
        col_norm = " ".join(str(col).upper().split())
        if col_norm in aliases_norm:
            renombrar[col] = aliases_norm[col_norm]

    if renombrar:
        df = df.rename(columns=renombrar)
    return df


# ------------------------------------------------------------------------------
# Validación de columnas
# ------------------------------------------------------------------------------
def validar_columnas_catalogo(df: pd.DataFrame) -> list:
    """Verifica que el catálogo tenga todas las columnas necesarias.
    Aplica la normalización de alias antes de validar."""
    df = _normalizar_columnas_catalogo(df)
    requeridas = [
        config.COL_SKU,
        config.COL_ITEM,
        config.COL_SUPPLIER,
        config.COL_SUPPLIER_ID,
    ] + config.DIVISIONES

    faltantes = [c for c in requeridas if c not in df.columns]
    return faltantes


def validar_columnas_bot(df: pd.DataFrame) -> list:
    """Verifica que el BOT tenga las columnas de SKU y proveedor."""
    requeridas = [config.COL_BOT_SKU, config.COL_BOT_SUPPLIER]
    return [c for c in requeridas if c not in df.columns]


def validar_columnas_bot_char(df: pd.DataFrame) -> list:
    """Verifica que el BOT Charcuterie tenga las columnas necesarias."""
    requeridas = [config.COL_BOT_CHAR_SUPPLIER, config.COL_BOT_CHAR_SKU, config.COL_BOT_CHAR_DESCR]
    return [c for c in requeridas if c not in df.columns]


# ------------------------------------------------------------------------------
# Procesamiento principal
# ------------------------------------------------------------------------------
def _construir_set_exp(df_exp: pd.DataFrame) -> set:
    """
    Retorna un set de (sku_norm, supplier_id_norm, division_canonica) extraído del
    archivo EXP. Cada tupla representa una combinación que debe ser obligatoriamente 'E'.
    Resuelve alias de división (Airlines, centerplate, Corporate services, etc.)
    usando config.ALIASES_COLUMNAS y config.DIVISIONES con búsqueda case-insensitive.
    """
    if df_exp is None or df_exp.empty:
        return set()

    alias_map = {}
    for div in config.DIVISIONES:
        alias_map[div.lower()] = div
    for alias, canonical in config.ALIASES_COLUMNAS.items():
        alias_map[alias.lower()] = canonical

    resultado = set()
    for _, row in df_exp.iterrows():
        sku = _normalizar_id(row.get(config.COL_EXP_SKU, ""))
        sid = _normalizar_id(row.get(config.COL_EXP_SUPPLIER_ID, ""))
        div_raw = str(row.get(config.COL_EXP_DIVISION, "")).strip()
        div_canon = alias_map.get(div_raw.lower(), div_raw)
        if sku and sid and div_canon:
            resultado.add((sku, sid, div_canon))
    return resultado


def procesar_catalogo(
    df_catalogo: pd.DataFrame,
    df_bot: pd.DataFrame,
    df_asignaciones: pd.DataFrame,
    df_exp=None,
    df_bot_char=None,
) -> pd.DataFrame:
    """
    Procesa el catálogo completo:
      1. Clasifica cada item en su categoría (usando IA para Local + carne)
      2. Calcula R/E para cada división según las reglas
      3. Compara con la clasificación previa de la persona (modelo de control)
      4. Une la información del distribuidor con la persona asignada
    """
    exp_set = _construir_set_exp(df_exp)

    df = df_catalogo.copy()

    # Renombrar columnas con alias conocidos (CENTERPLATE -> SPORTS & LEISURE, etc.)
    df = _normalizar_columnas_catalogo(df)

    # --- Set de SKUs presentes en el BOT (como strings normalizados) ---
    if not df_bot.empty and config.COL_BOT_SKU in df_bot.columns:
        skus_bot = set(df_bot[config.COL_BOT_SKU].apply(clasificador._normalizar_id))
        skus_bot.discard("")  # quitar entradas vacías si las hay
    else:
        skus_bot = set()

    # --- Set de SKUs presentes en el BOT Charcuterie ---
    if df_bot_char is not None and not df_bot_char.empty and config.COL_BOT_CHAR_SKU in df_bot_char.columns:
        skus_bot_char = set(df_bot_char[config.COL_BOT_CHAR_SKU].apply(clasificador._normalizar_id))
        skus_bot_char.discard("")
    else:
        skus_bot_char = set()

    # --- Pre-detectar items con "Local" para llamar a la IA en lote ---
    items = df[config.COL_ITEM].fillna("").astype(str).tolist()
    items_local = [it for it in items if clasificador.es_local(it)
                   and not clasificador.es_banned(it)]
    items_local_unicos = list(set(items_local))

    cache_carne = {}
    if items_local_unicos:
        cache_carne = ia_carne.precalentar_cache(items_local_unicos)

    def es_carne_func(item: str) -> bool:
        return cache_carne.get(item, False)

    # --- Clasificar cada fila (con caché para items repetidos) ---
    categorias = []
    re_por_div = {div: [] for div in config.DIVISIONES}

    total_filas = len(df)
    progreso_clas = st.progress(0, text=f"Clasificando productos: 0/{total_filas:,}...")

    # Convertir el iterador a algo más rápido: extraer las 3 columnas como listas
    items_lista = df[config.COL_ITEM].astype(str).fillna("").tolist()
    skus_lista = df[config.COL_SKU].tolist()
    sids_lista = df[config.COL_SUPPLIER_ID].tolist()

    # CACHÉ: la clasificación depende de (item, supplier_id) — items idénticos
    # con el mismo supplier_id dan el mismo resultado. Esto evita reprocesar
    # miles de items repetidos.
    cache_clasificacion = {}    # {(item, sid_norm): categoria}
    cache_re_por_cat = {}        # {categoria: re_dict_estatico}  (para cats sin BOT)
    cache_re_initial = {}        # {(sku_norm,): re_dict}  (Initial Catalog usa BOT)

    # Pre-normalizar SKUs y SIDs (evita recalcular en cada llamada)
    skus_norm_lista = [clasificador._normalizar_id(s) for s in skus_lista]
    sids_norm_lista = [clasificador._normalizar_id(s) for s in sids_lista]

    PASO_PROGRESO = max(1, total_filas // 100)

    for idx in range(total_filas):
        item = items_lista[idx]
        sku_norm = skus_norm_lista[idx]
        sid_norm = sids_norm_lista[idx]

        # 1) Clasificar (con caché)
        clave_cat = (item, sid_norm)
        cat = cache_clasificacion.get(clave_cat)
        if cat is None:
            cat = clasificador.clasificar_categoria(
                item, es_carne_func, supplier_id=sid_norm
            )
            cache_clasificacion[clave_cat] = cat
        categorias.append(cat)

        # 2) Calcular R/E (con caché diferenciado)
        # Para Initial Catalog depende del SKU (puede estar o no en BOT) +
        # del item (por las reglas específicas tipo SCHOOL SERVICES).
        # Para las demás categorías depende de la categoría + item + supplier
        # (las reglas por supplier pueden forzar todo a R).
        if cat == config.CAT_INITIAL_CATALOG:
            clave_re = (cat, sku_norm, item, sid_norm)
        else:
            clave_re = (cat, item, sid_norm)

        re_dict = cache_re_por_cat.get(clave_re)
        if re_dict is None:
            re_dict = clasificador.calcular_re_por_division(
                cat, sku_norm, skus_bot, item=item, supplier_id=sid_norm
            )
            cache_re_por_cat[clave_re] = re_dict

        for div in config.DIVISIONES:
            re_por_div[div].append(re_dict[div])

        if cat == config.CAT_NON_CONTRACTED and sku_norm in skus_bot_char:
            for div in config.DIVISIONES_BOT_CHARCUTERIE:
                re_por_div[div][-1] = "E"

        # Actualizar barra cada PASO_PROGRESO filas
        if (idx + 1) % PASO_PROGRESO == 0 or (idx + 1) == total_filas:
            progreso_clas.progress(
                (idx + 1) / total_filas,
                text=f"Clasificando productos: {idx + 1:,}/{total_filas:,}..."
            )

    progreso_clas.empty()

    # Para el estado de divisiones BOT: set de (sku, supplier) con normalización robusta.
    # Un item está en el BOT solo si coinciden AMBOS: SKU y nombre del proveedor.
    _tiene_cols_bot = (
        not df_bot.empty
        and config.COL_BOT_SKU in df_bot.columns
        and config.COL_BOT_SUPPLIER in df_bot.columns
    )
    if _tiene_cols_bot:
        _bot_skus   = df_bot[config.COL_BOT_SKU].apply(_normalizar_id)
        _bot_sups   = df_bot[config.COL_BOT_SUPPLIER].fillna("").astype(str).str.strip().str.upper()
        _bot_set    = set(zip(_bot_skus, _bot_sups))
        _bot_set.discard(("", ""))
    else:
        _bot_set = set()

    _cat_sups = df[config.COL_SUPPLIER].fillna("").astype(str).str.strip().str.upper().tolist()
    en_bot_arr = [
        (_normalizar_id(sku), sup) in _bot_set
        for sku, sup in zip(skus_lista, _cat_sups)
    ]

    # Mostrar info de caché
    items_unicos = len(cache_clasificacion)
    if items_unicos < total_filas:
        ahorro = (1 - items_unicos / total_filas) * 100
        st.caption(
            f"⚡ Optimización: {items_unicos:,} items únicos procesados "
            f"de {total_filas:,} totales ({ahorro:.1f}% ahorrado por caché)."
        )

    df["Categoria"] = categorias

    # --- Crear columnas calculadas (R/E según reglas) ---
    for div in config.DIVISIONES:
        df[f"{div} (calculado)"] = re_por_div[div]

    # --- Modelo de control: 5 estados por celda ---
    #   "(estado)": "ok"        -> persona acertó (verde)
    #               "error"     -> persona se equivocó (rojo)
    #               "blank"     -> celda vacía (gris con guion)
    #               "invalid"   -> item con descripción inválida (naranja, no se compara)
    #               "bot_error" -> error en Initial Catalog según BOT (naranja oscuro)
    # NOTA: En invalid se preserva el valor original que la persona puso.
    for div in config.DIVISIONES:
        previo_raw = df[div]
        es_blank = previo_raw.isna() | (previo_raw.astype(str).str.strip() == "") | \
                   (previo_raw.astype(str).str.lower().isin(["nan", "none"]))
        previo = previo_raw.astype(str).str.strip().str.upper()
        calculado = df[f"{div} (calculado)"].astype(str).str.strip().str.upper()

        # Detectar items inválidos por categoría
        es_invalid = df["Categoria"] == config.CAT_REVISAR

        # Construir el estado
        estado = ["ok" if c else "error" for c in (previo == calculado)]
        estado = ["blank" if b else e for b, e in zip(es_blank, estado)]
        estado = ["invalid" if inv else e for inv, e in zip(es_invalid, estado)]
        df[f"{div} (estado)"] = estado

        # Para Initial Catalog en divisiones BOT: recalcular estado directo contra BOT
        # (evita que un en_bot incorrecto en la clasificación afecte el color)
        if div in config.DIVISIONES_REQUIEREN_BOT:
            es_ic_list = (df["Categoria"] == config.CAT_INITIAL_CATALOG).tolist()
            previo_list = previo.tolist()
            for i in range(len(estado)):
                if not es_ic_list[i] or estado[i] in ("blank", "invalid"):
                    continue
                if en_bot_arr[i] and previo_list[i] == "E":
                    estado[i] = "ok"
                elif en_bot_arr[i]:
                    estado[i] = "bot_error"
                elif previo_list[i] == "R":
                    estado[i] = "ok"
                else:
                    estado[i] = "bot_error"
            df[f"{div} (estado)"] = estado

        # Reemplazar visualmente los blanks por un guion (solo en blanks NO inválidos)
        df[div] = df[div].astype(object)
        mostrar_guion = es_blank & ~es_invalid
        df.loc[mostrar_guion, div] = "—"

    # --- Validación cruzada con archivo EXP ---
    # EXP es autoridad: los ítems listados allí DEBEN ser "E".
    # Usamos el estado ya calculado como proxy del valor que puso el usuario:
    #   estado="error" + calculado="R"  →  usuario puso "E" (correcto per EXP)  → ok
    #   estado="ok"    + calculado="R"  →  usuario puso "R" (incorrecto per EXP) → error_exp
    if exp_set:
        skus_norm = df[config.COL_SKU].apply(_normalizar_id)
        sids_norm = df[config.COL_SUPPLIER_ID].apply(_normalizar_id)
        for div in config.DIVISIONES:
            col_calc   = f"{div} (calculado)"
            col_estado = f"{div} (estado)"
            if col_calc not in df.columns:
                continue
            for idx in df.index:
                clave = (skus_norm[idx], sids_norm[idx], div)
                if clave not in exp_set:
                    continue
                if df.at[idx, col_calc] != "R":
                    continue
                est = df.at[idx, col_estado]
                if est == "error":
                    # Usuario puso "E" → EXP lo confirma → correcto
                    df.at[idx, col_estado] = "ok"
                elif est == "ok":
                    # Usuario puso "R" → EXP dice que debe ser "E" → error
                    df.at[idx, col_estado] = "error_exp"
                # blank e invalid se dejan sin cambios

    # --- Unir con asignaciones (match por Supplier ID) ---
    # El catálogo ya tiene Supplier Name, pero hacemos el match por ID
    # porque es más confiable. La persona asignada se trae de asignaciones.
    if not df_asignaciones.empty:
        # Normalizar IDs: maneja floats (1234.0) y espacios
        df["_sid_match"] = df[config.COL_SUPPLIER_ID].apply(_normalizar_id)
        df_asign_norm = df_asignaciones.copy()
        df_asign_norm["_sid_match"] = df_asign_norm[config.COL_ASSIGN_SUPPLIER_ID].apply(_normalizar_id)

        # Quitar filas duplicadas en asignaciones (por si hay 2 entradas con mismo ID)
        df_asign_norm = df_asign_norm.drop_duplicates(subset=["_sid_match"], keep="first")

        # Solo traer la persona; el Supplier Name del catálogo ya existe
        df = df.merge(
            df_asign_norm[["_sid_match", config.COL_ASSIGN_PERSON]],
            how="left",
            on="_sid_match",
        )

        # Diagnóstico: cuántas filas quedaron sin asignar
        sin_asignar = df[config.COL_ASSIGN_PERSON].isna().sum()
        if sin_asignar > 0:
            ids_sin_match = (
                df[df[config.COL_ASSIGN_PERSON].isna()]["_sid_match"]
                .unique()[:5]
                .tolist()
            )
            st.warning(
                f"⚠️ {sin_asignar} filas no encontraron persona asignada. "
                f"Primeros IDs sin match: {ids_sin_match}. "
                f"Verifica que esos IDs existan en data/asignaciones.xlsx."
            )

        df.drop(columns=["_sid_match"], inplace=True)
    else:
        df[config.COL_ASSIGN_PERSON] = None

    return df


# ------------------------------------------------------------------------------
# Construcción de la tabla "vista" (ordenada para mostrar al usuario)
# ------------------------------------------------------------------------------
def construir_tabla_vista(df: pd.DataFrame) -> pd.DataFrame:
    """
    Devuelve un DataFrame con las columnas en el ORDEN solicitado:
    Persona | Distribuidor | Unidad | SKU | Producto | Categoría | 7 divisiones
    """
    columnas_base = [
        config.COL_ASSIGN_PERSON,
        config.COL_SUPPLIER,
        config.COL_SUPPLIER_ID,
        config.COL_SKU,
        config.COL_ITEM,
        "Categoria",
    ]
    # Para cada división mostramos el VALOR PREVIO (lo que puso la persona)
    # y el modelo de control se aplica con color (verde/rojo)
    columnas_div = list(config.DIVISIONES)
    cols = columnas_base + columnas_div
    cols = [c for c in cols if c in df.columns]
    return df[cols].copy()


# ------------------------------------------------------------------------------
# Exportar a Excel coloreado (para descarga)
# ------------------------------------------------------------------------------
def exportar_excel_coloreado(df_completo: pd.DataFrame) -> bytes:
    """
    Genera un Excel con las columnas R/E coloreadas según el modelo de control:
      - VERDE: persona acertó
      - ROJO : persona se equivocó
      - GRIS : celda vacía (faltó completar)
    Devuelve los bytes del archivo.
    """
    df_vista = construir_tabla_vista(df_completo)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_vista.to_excel(writer, index=False, sheet_name="Catalogo")
        wb = writer.book
        ws = writer.sheets["Catalogo"]

        # Estilos
        font_blanco = Font(color="FFFFFF", bold=True)
        fill_ok       = PatternFill("solid", fgColor=config.COLOR_OK.lstrip("#"))
        fill_err      = PatternFill("solid", fgColor=config.COLOR_ERROR.lstrip("#"))
        fill_blank    = PatternFill("solid", fgColor=config.COLOR_BLANK.lstrip("#"))
        fill_invalid  = PatternFill("solid", fgColor=config.COLOR_INVALID.lstrip("#"))
        fill_bot_err  = PatternFill("solid", fgColor=config.COLOR_BOT_MISMATCH.lstrip("#"))
        mapa_fill = {
            "ok": fill_ok,
            "error": fill_err,
            "blank": fill_blank,
            "invalid": fill_invalid,
            "error_exp": fill_err,
            "bot_error": fill_bot_err,
        }

        headers = [c.value for c in ws[1]]
        for div in config.DIVISIONES:
            if div not in headers:
                continue
            col_idx = headers.index(div) + 1
            col_letra = get_column_letter(col_idx)

            estados = df_completo[f"{div} (estado)"].tolist()
            for fila_idx, estado in enumerate(estados, start=2):
                celda = ws[f"{col_letra}{fila_idx}"]
                fill = mapa_fill.get(estado)
                if fill is not None:
                    celda.fill = fill
                    celda.font = font_blanco
                celda.alignment = celda.alignment.copy(horizontal="center")

        # Auto-ajustar ancho de columnas
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None),
                          default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    return output.getvalue()


# ------------------------------------------------------------------------------
# Función para Streamlit: aplicar colores en la tabla en pantalla
# ------------------------------------------------------------------------------
def estilizar_tabla(df_vista: pd.DataFrame, df_completo: pd.DataFrame):
    """
    Devuelve un Styler de pandas con las celdas R/E coloreadas según 3 estados:
      - "ok"    -> verde
      - "error" -> rojo
      - "blank" -> gris (la persona no llenó la celda)
    """
    cols_div = [c for c in df_vista.columns if c in config.DIVISIONES]
    df_completo_reset = df_completo.reset_index(drop=True)

    estilo_ok = (
        f"background-color: {config.COLOR_OK}; color: white; "
        "font-weight: bold; text-align: center;"
    )
    estilo_err = (
        f"background-color: {config.COLOR_ERROR}; color: white; "
        "font-weight: bold; text-align: center;"
    )
    estilo_blank = (
        f"background-color: {config.COLOR_BLANK}; color: white; "
        "font-weight: bold; text-align: center;"
    )
    estilo_invalid = (
        f"background-color: {config.COLOR_INVALID}; color: white; "
        "font-weight: bold; text-align: center;"
    )
    estilo_bot_err = (
        f"background-color: {config.COLOR_BOT_MISMATCH}; color: white; "
        "font-weight: bold; text-align: center;"
    )
    mapa = {
        "ok": estilo_ok,
        "error": estilo_err,
        "blank": estilo_blank,
        "invalid": estilo_invalid,
        "error_exp": estilo_err,
        "bot_error": estilo_bot_err,
    }

    def aplicar_estilos(df):
        estilos = pd.DataFrame("", index=df.index, columns=df.columns)
        for col in cols_div:
            estado_col = f"{col} (estado)"
            if estado_col not in df_completo_reset.columns:
                continue
            estados = df_completo_reset[estado_col].iloc[: len(df)].values
            estilos[col] = [mapa.get(e, "") for e in estados]
        return estilos

    return df_vista.style.apply(aplicar_estilos, axis=None)
